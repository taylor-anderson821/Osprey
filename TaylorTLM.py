ALTITUDE_MOVING_AVG_WINDOW = 5  # Number of previous samples to include in altitude moving average
THERMAL_CANDIDATE_WINDOW = 30 # the number of points to look forward or back for thermal peak candidate 
THERMAL_PEAK_MIN_ALTITUDE = 100 # minimum altitude of a thermal peak to be considered
THERMAL_MINIMUM_THRESHOLD = 25 #minimum altitude gain before recognizing a thermal
THERMAL_MAX_CLIMB_RATE_FPS = 9 # maximum thermal climb rate (ft/s) to consider for thermal peak candidates (filters out launches)
THERMAL_MAX_CLIMB_RATE_LOOKBACK = 5 #number of seconds to look back to check for high climb rate before thermal peak (all samples in seconds)
LAUNCH_MIN_CLIMB_RATE_FPS = 21 # minimum launch climb rate (ft/s) to consider for launch peak candidates
LAUNCH_MIN_CLIMB_RATE_LOOKBACK = 5 #number of seconds to look back to check for high climb rate before launch peak (all samples in seconds)
TROUGH_CANDIDATE_WINDOW = 10 # number of seconds to look back to check for low climb rate before trough bottom candidate
SESSION_MINIMUM_DURATION = 30

import struct
import numpy
import os
import openpyxl
import datetime

class VarioRecord:
        def __init__(self,timestamp, altitude, climb_rate):
            self.timestamp = timestamp
            self.altitude = altitude
            self.climb_rate = climb_rate
            self.altitude_smoothed = 0.0
            self.thermal_peak_candidate = False
            self.thermal_peak = False
            self.launch_peak_candidate = False
            self.launch_peak = False
            self.trough_bottom_candidate = False
            self.trough_bottom = False
    
class ThermalRecord:
    def __init__(self, session_number, thermal_number, start_time, start_index, end_time, end_index, start_altitude, end_altitude, duration, altitude_gain,
                    avg_climb_rate):
        self.session_number = session_number
        self.thermal_number = thermal_number
        self.start_time = start_time
        self.start_index = start_index
        self.end_time = end_time
        self.end_index = end_index
        self.start_altitude = start_altitude
        self.end_altitude = end_altitude
        self.duration = duration
        self.altitude_gain = altitude_gain
        self.avg_climb_rate = avg_climb_rate

class SessionRecord:
    def __init__(self, start_datetime, start_date):
        self.start_time = start_datetime
        self.start_date = start_date
        self.launch_count = 0
        self.thermal_count = 0
        self.total_thermal_altitude_gain = 0
        self.total_thermal_duration = 0
        self.total_thermal_duration_hms = 0
        self.session_duration = 0
        self.session_duration_hms = ''
        self.thermal_launch_ratio = 0.0

class DailySessionSummary:
    def __init__(self, date):
        self.date = date
        self.launch_count = 0
        self.thermal_count = 0
        self.total_thermal_altitude_gain = 0
        self.total_thermal_duration = 0
        self.total_thermal_duration_hms = 0
        self.session_duration = 0
        self.session_duration_hms = 0
        self.thermal_launch_ratio = 0.0
        self.total_thermal_duration_pct = 0.0

tlm_file_path = ''

def process_header_packets(f):
    """
    Reads first 4 bytes as timestamp and checks if it's a header packet.
    If timestamp is 0xFFFFFFFF, skips 32 bytes and recursively checks again.
    Otherwise, calls process_payload_packet.
    """
    # Read first 4 bytes as timestamp
    while True:
        timestamp_bytes = f.read(4)
    
        # Check if we've reached end of file
        if len(timestamp_bytes) < 4:
            print("End of file reached.")
            return
        
        # Convert to unsigned 32-bit integer (little-endian)
        timestamp = struct.unpack('<I', timestamp_bytes)[0]
                
        # Check if timestamp is 0xFFFFFFFF
        if timestamp == 0xFFFFFFFF:
            skipped = f.read(32)
            if len(skipped) < 32:
                print("Not enough bytes to skip. End of file.")
                return
        else:
            return

def summarize_daily_sessions(session_records):

    daily_session_summary = []
    j = -1
    current_session_date = datetime.datetime.now() # initialize variable to a datetime
    global tlm_file_path 

    for i in range(0, len(session_records)):
        if(session_records[i].start_date != current_session_date):
            j += 1 # index for the day we are generating a record for
            current_session_date = session_records[i].start_date
            daily_session_summary.append(DailySessionSummary(current_session_date))

        daily_session_summary[j].launch_count += session_records[i].launch_count
        daily_session_summary[j].thermal_count += session_records[i].thermal_count
        daily_session_summary[j].total_thermal_altitude_gain += session_records[i].total_thermal_altitude_gain
       
        daily_session_summary[j].total_thermal_duration += session_records[i].total_thermal_duration
        total_thermal_duration_hms = str(datetime.timedelta(seconds= int(daily_session_summary[j].total_thermal_duration)))
        daily_session_summary[j].total_thermal_duration_hms = total_thermal_duration_hms
        
        daily_session_summary[j].session_duration += session_records[i].session_duration
        total_session_duration_hms = str(datetime.timedelta(seconds= int(daily_session_summary[j].session_duration)) )
        daily_session_summary[j].session_duration_hms = total_session_duration_hms
        if (daily_session_summary[j].session_duration != 0):
            daily_session_summary[j].total_thermal_duration_pct = daily_session_summary[j].total_thermal_duration  / daily_session_summary[j].session_duration

        if (daily_session_summary[j].launch_count != 0):
            daily_session_summary[j].thermal_launch_ratio = daily_session_summary[j].thermal_count / daily_session_summary[j].launch_count

    daily_flight_summary_path = os.path.splitext(tlm_file_path)[0] + '_daily_summary.xlsx'
    
    try:
        with open(daily_flight_summary_path, 'r') as file:
            summary_file_exists = True
    except FileNotFoundError:
        summary_file_exists = False

    if summary_file_exists == False: # create the summary file
        wb = openpyxl.Workbook()
        ws = wb["Sheet"]
        ws.title = "Daily Summary"
        headers = ['Date', 'Launches', 'Thermals', 'Total Thm Gain (ft)', 'Total Thm Duration', 'Total Flying Duration', 'Thermal/Lauch Ratio', 'Thermal Duration Pct' ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        columns = ['A','B','C','D','E','F','G','H']
        for column_letter in columns:
            ws["{}1".format(column_letter)].font = openpyxl.styles.Font(bold=True)
            ws["{}1".format(column_letter)].alignment = openpyxl.styles.Alignment(horizontal = 'center', wrap_text = True)
    
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        startRow = 2
        
    else:
        #Open the existing summary file
        wb = openpyxl.load_workbook(filename = daily_flight_summary_path)
        #Get the current Active Sheet
        ws = wb['Daily Summary']
        startRow = ws.max_row + 1
    
    for i in range(0, len(daily_session_summary)):
        rowIndex = i + startRow
        ws.cell(row=rowIndex, column=1, value=daily_session_summary[i].date)
        ws.cell(row=rowIndex, column=2, value=daily_session_summary[i].launch_count)
        ws.cell(row=rowIndex, column=3, value=daily_session_summary[i].thermal_count)
        cell = ws.cell(row=rowIndex, column=4, value=daily_session_summary[i].total_thermal_altitude_gain)
        cell.number_format = '#,##0'
        ws.cell(row=rowIndex, column=5, value=daily_session_summary[i].total_thermal_duration_hms)
        ws["E{}".format(rowIndex)].alignment = openpyxl.styles.Alignment(horizontal='right')
        ws.cell(row=rowIndex, column=6, value=daily_session_summary[i].session_duration_hms)
        ws["F{}".format(rowIndex)].alignment = openpyxl.styles.Alignment(horizontal='right')
        cell = ws.cell(row=rowIndex, column=7, value=round(daily_session_summary[i].thermal_launch_ratio, 2))
        cell.number_format = '0.00'
        cell = ws.cell(row=rowIndex, column=8, value=round(daily_session_summary[i].total_thermal_duration_pct, 2))
        cell.number_format = '0%'

    wb.save(daily_flight_summary_path)  
    print('Daily flight summary saved to:', daily_flight_summary_path)


def smooth_altitude_readings(vario_records):
    #Apply simple moving average to altitude data for current session
    
    average = 0.0

    for i in range(0, len(vario_records)):
        if(i < ALTITUDE_MOVING_AVG_WINDOW):
            vario_records[i].altitude_smoothed = vario_records[i].altitude #i-ALTITUDE_MOVING_AVG_WINDOW):i
        else:
            average = numpy.sum([vario_records[j].altitude for j in range(i - ALTITUDE_MOVING_AVG_WINDOW+1, i+1)]) / ALTITUDE_MOVING_AVG_WINDOW 
            vario_records[i].altitude_smoothed = float(average)

def identify_thermal_peaks(vario_records, session_records, session_number):
    # Identifies thermal peaks in the vario_records for the current session using the following logic:
    # 1. Identify thermal peak candidates based on altitude comparisons (higher than surrounding points)
    # 2. From candidates, select the highest point in each cluster as a thermal peak
    # 3. Disqualify thermal peaks associated with high climb rates (likely launches)

    # identify any thermal peak candidates that are higher than the surrounding +/- THERMAL_CANDIDATE_WINDOW records
    for i in range(THERMAL_CANDIDATE_WINDOW, len(vario_records)-THERMAL_CANDIDATE_WINDOW):
        if(vario_records[i].altitude_smoothed > vario_records[i - THERMAL_CANDIDATE_WINDOW].altitude_smoothed and 
           vario_records[i].altitude_smoothed > vario_records[i + THERMAL_CANDIDATE_WINDOW].altitude_smoothed and
           vario_records[i].altitude_smoothed > THERMAL_PEAK_MIN_ALTITUDE):
               vario_records[i].thermal_peak_candidate = True

    # candidates will be in clusters -- only keep the highest point in each cluster   
    # search for records marked as thermal_peak_candidate and see if they are the highest point in +/- THERMAL_CANDIDATE_WINDOW 
    for i in range(THERMAL_CANDIDATE_WINDOW, len(vario_records)-THERMAL_CANDIDATE_WINDOW):
        # look for thermal peak candidates
        if(vario_records[i].thermal_peak_candidate == True):
            maximum_thermal_peak = 0.0
            #find the maximum altitude in the candidate cluster
            for j in range(i - THERMAL_CANDIDATE_WINDOW, i+THERMAL_CANDIDATE_WINDOW):
                if(vario_records[j].altitude_smoothed > maximum_thermal_peak):
                    maximum_thermal_peak = vario_records[j].altitude_smoothed

            #identify the record(s) that match the maximum altitude and mark as thermal peak
            if(vario_records[i].altitude_smoothed == maximum_thermal_peak):
                vario_records[i].thermal_peak = True 
                #confirm that this is not a thermal associated with a powered launch (high climb rate)
                #look at increase in altitue over THERMAL_MAX_CLIMB_RATE_LOOKBACK seconds and if
                # it exceeds max climb rate, disqualify as thermal peak
                if(vario_records[i].altitude_smoothed - vario_records[i - THERMAL_MAX_CLIMB_RATE_LOOKBACK].altitude_smoothed
                   > THERMAL_MAX_CLIMB_RATE_FPS * THERMAL_MAX_CLIMB_RATE_LOOKBACK):
                    vario_records[i].thermal_peak = False
                if(vario_records[i].thermal_peak == True):
                    print("session", session_number, ": Thermal peak identified at", vario_records[i].timestamp, "s, Altitude:", 
                      round(vario_records[i].altitude_smoothed,2), "ft, Vario:", round(vario_records[i].climb_rate,2), "ft/min")      
 
def identify_launch_peaks(vario_records, session_records, session_number):
# Identifies launch peaks in the vario_records for the current session using the following logic:
# 1. Identify launch peak candidates based on meeting a minium climb rate during the lookback period
# 2. Candidates will be in clusters, mark the climb peak after the last candidate in the cluster

    for i in range(LAUNCH_MIN_CLIMB_RATE_LOOKBACK, len(vario_records)-LAUNCH_MIN_CLIMB_RATE_LOOKBACK):
        # identify for launch peak candidates
        maximum_launch_peak = 0
        # mark any record where the climb rate exceeded the threshold
        for j in range(i - LAUNCH_MIN_CLIMB_RATE_LOOKBACK, i):
                if(vario_records[j].climb_rate > LAUNCH_MIN_CLIMB_RATE_FPS):
                    maximum_launch_peak = vario_records[j].climb_rate
        if(maximum_launch_peak > LAUNCH_MIN_CLIMB_RATE_FPS):
                vario_records[i].launch_peak_candidate = True

    for i in range(LAUNCH_MIN_CLIMB_RATE_LOOKBACK, len(vario_records)-LAUNCH_MIN_CLIMB_RATE_LOOKBACK):
        if(vario_records[i].launch_peak_candidate == False and vario_records[i-1].launch_peak_candidate == True):
            # mark the launch peak at this record
            vario_records[i].launch_peak = True
            session_records[session_number].launch_count += 1
            print("session", session_number, ": Launch peak identified at index", i, vario_records[i].timestamp, "s, "
            "Altitude:", round(vario_records[i].altitude_smoothed,2))

def identify_trough_bottoms(vario_records, session_number):
# Identifies trough bottoms in the vario_records for the current session using the following logic:
# 1. Identify trough bottom candidates based on altitude comparisons (lower than surrounding points)
# 2. From candidates, select the lowest point in each cluster as a trough bottom

    # identify any trough bottom candidates that are lower than the surrounding +/- TROUGH_CANDIDATE_WINDOW records
    for i in range(TROUGH_CANDIDATE_WINDOW, len(vario_records)-TROUGH_CANDIDATE_WINDOW):
        if(vario_records[i].altitude_smoothed < vario_records[i - TROUGH_CANDIDATE_WINDOW].altitude_smoothed and 
           vario_records[i].altitude_smoothed < vario_records[i + TROUGH_CANDIDATE_WINDOW].altitude_smoothed):
               vario_records[i].trough_bottom_candidate = True

    # candidates will be in clusters -- only keep point with the highest altitude in each cluster   
    # search for records marked as trough_bottom_candidate and see if they are the lowest point in +/- TROUGH_CANDIDATE_WINDOW
    for i in range(TROUGH_CANDIDATE_WINDOW, len(vario_records)-TROUGH_CANDIDATE_WINDOW):
        # look for trough bottom candidates
        if(vario_records[i].trough_bottom_candidate == True):
            minimum_trough_bottom = 10000  # set an arbitrary high value to work down from
            #find the minimum altitude in the candidate cluster
            for j in range(i - TROUGH_CANDIDATE_WINDOW, i+TROUGH_CANDIDATE_WINDOW):
                if(vario_records[j].altitude_smoothed < minimum_trough_bottom):
                    minimum_trough_bottom = vario_records[j].altitude_smoothed

            #identify the record(s) that match the minimum altitude and mark as trough bottom
            if(vario_records[i].altitude_smoothed == minimum_trough_bottom):
                    vario_records[i].trough_bottom = True 
                    # print("session", session_number, ": Trough bottom identified at index", i, " timestamp", vario_records[i].timestamp, "Altitude:",vario_records[j].altitude_smoothed )
   

def identify_caught_thermals(vario_records, session_records, session_number, thermal_records, wb):
# Identifies caught thermals.  Process is:  
# 1. Find each thermal peak in a session
# 2. Work backwards to find the next trough bottom or launch peak
# 3. If a launch peak is the preceding event before a thermal peak, use it's altitude as the start of the thermal, if a trough bottom is found
# work backwards to the next event. If it is a higher trough bottom or a launch peak, use previous trough bottom as start of thermal.   
    
    # identify for launch peak candidates
    thermal_index = 0
    last_trough_bottom_altitude = 0.0
    last_trough_bottom_timestamp = 0.0
    last_trough_bottom_index = 0
    i = len(vario_records) - 1
    j = 0
    sheet_name = "Ssn" + str(session_number)
    ws = wb[sheet_name]

     # create a static variable to track session number
    if not hasattr(identify_caught_thermals, "thermal_number"):
        process_payload_packet.thermal_number = 0 

    session_duration = vario_records[len(vario_records) - 1].timestamp 
    session_duration_hms = str(datetime.timedelta(seconds=int(session_duration)))
    session_records[session_number].session_duration = session_duration
    session_records[session_number].session_duration_hms = session_duration_hms

    # work backwards in time from the end of the session
    while(i>0):
        if(vario_records[i].thermal_peak == True): # found a thermal peak, now work backwards to find preceding event(s)  
            #last_thermal_peak_index = i
            last_event_was_a_trough = False
            j = i
            while(j > 0):                    
                j -= 1
                if(vario_records[j].launch_peak == True and last_event_was_a_trough == False):
                   # mark caught thermal from launch peak to thermal peak
                    start_time = round(vario_records[j].timestamp, 1)
                    end_time = round(vario_records[i].timestamp, 1)
                    start_altitude = round(vario_records[j].altitude_smoothed, 1)
                    end_altitude = round(vario_records[i].altitude_smoothed, 1)
                    duration = round(end_time - start_time, 1)
                    altitude_gain = round(end_altitude - start_altitude, 1)
                    avg_climb_rate = round(altitude_gain/duration, 1)
                    if(altitude_gain >= THERMAL_MINIMUM_THRESHOLD):
                        thermal_records.append(ThermalRecord(session_number, thermal_index, start_time, j, end_time, i, start_altitude, end_altitude, duration, altitude_gain, avg_climb_rate))
                        ws.cell(row=j, column=4, value=start_altitude) #mark thermal beginning in session sheet in XLS
                        ws.cell(row=i, column=5, value=end_altitude) #mark thermal end in session sheet in XLS
                        session_records[session_number].thermal_count += 1
                        session_records[session_number].total_thermal_altitude_gain += altitude_gain
                        session_records[session_number].total_thermal_duration += duration
                        print("Session", session_number, "Thermal", thermal_index, "- Start alt", round(start_altitude,1), "\tEnd alt", round(end_altitude,1)) #, "start index", last_trough_bottom_index, "end index", i)
                    i=j # move starting point for next thermal peak to the start of the last thermal
                    break

                elif(vario_records[j].trough_bottom == True and last_event_was_a_trough == True):
                    #print("current trough ", vario_records[j].altitude_smoothed, "last trough", last_trough_bottom_altitude)
                    if(vario_records[j].altitude_smoothed > last_trough_bottom_altitude):
                        start_time = round(last_trough_bottom_timestamp, 1)
                        end_time = round(vario_records[i].timestamp, 1)
                        start_altitude = round(last_trough_bottom_altitude, 1)
                        end_altitude = round(vario_records[i].altitude_smoothed, 1)
                        duration = round(end_time - start_time, 1)
                        altitude_gain = round(end_altitude - start_altitude, 1)
                        avg_climb_rate = round(altitude_gain/duration, 1)
                        if(altitude_gain >= THERMAL_MINIMUM_THRESHOLD):
                            thermal_records.append(ThermalRecord(session_number, thermal_index, start_time, last_trough_bottom_index, 
                                                             end_time, i, start_altitude, end_altitude, duration, altitude_gain, avg_climb_rate ))                       
                            ws.cell(row=last_trough_bottom_index, column=4, value=start_altitude) #mark thermal beginning in session sheet in XLS
                            ws.cell(row=i, column=5, value=end_altitude) #mark thermal end in session sheet in XLS
                            session_records[session_number].thermal_count += 1
                            session_records[session_number].total_thermal_altitude_gain += altitude_gain
                            session_records[session_number].total_thermal_duration += duration
                            print("Session", session_number, "Thermal", thermal_index, "- Start alt", round(start_altitude,1), "\tEnd alt", round(end_altitude,1)) #, "start index", last_trough_bottom_index, "end index", i)
                        i=j # move starting point for next thermal peak to the start of the last thermal
                        break
                    else:
                        last_trough_bottom_altitude = vario_records[j].altitude_smoothed
                        last_trough_bottom_timestamp = vario_records[j].timestamp
                        last_trough_bottom_index = j

                elif(vario_records[j].thermal_peak == True and last_event_was_a_trough == True):
                    # mark caught thermal from the previously processed trough to thermal peak
                    start_time = round(last_trough_bottom_timestamp, 1)
                    end_time = round(vario_records[i].timestamp, 1)
                    start_altitude = round(last_trough_bottom_altitude, 1)
                    end_altitude = round(vario_records[i].altitude_smoothed, 1)
                    duration = round(end_time - start_time, 1)
                    altitude_gain = round(end_altitude - start_altitude, 1)
                    avg_climb_rate = round(altitude_gain/duration, 1)
                    if(altitude_gain >= THERMAL_MINIMUM_THRESHOLD):
                        thermal_records.append(ThermalRecord(session_number, thermal_index, start_time, last_trough_bottom_index, end_time, i, start_altitude, end_altitude, duration, altitude_gain, avg_climb_rate))
                        ws.cell(row=last_trough_bottom_index, column=4, value=start_altitude) #mark thermal beginning in session sheet in XLS
                        ws.cell(row=i, column=5, value=end_altitude) #mark thermal end in session sheet in XLS
                        session_records[session_number].thermal_count += 1
                        session_records[session_number].total_thermal_altitude_gain += altitude_gain
                        session_records[session_number].total_thermal_duration += duration
                        print("Session", session_number, "Thermal", thermal_index, "- Start alt", round(start_altitude,1), "\tEnd alt", round(end_altitude,1)) #, "start index", last_trough_bottom_index, "end index", i)
                    i=j+1 # since we're on a thermal peak, move starting point +1 so that decrement at start of loop will start us on this peak
                    break

                elif(vario_records[j].launch_peak == True and last_event_was_a_trough == True):
                    # mark caught thermal from the previously processed trough to thermal peak
                    start_time = round(last_trough_bottom_timestamp, 1)
                    end_time = round(vario_records[i].timestamp, 1)
                    start_altitude = round(last_trough_bottom_altitude, 1)
                    end_altitude = round(vario_records[i].altitude_smoothed, 1)
                    duration = round(end_time - start_time, 1)
                    altitude_gain = round(end_altitude - start_altitude, 1)
                    avg_climb_rate = round(altitude_gain/duration, 1)
                    if(altitude_gain >= THERMAL_MINIMUM_THRESHOLD):
                        thermal_records.append(ThermalRecord(session_number, thermal_index, start_time, last_trough_bottom_index, end_time, i, start_altitude, end_altitude, duration, altitude_gain, avg_climb_rate))
                        ws.cell(row=last_trough_bottom_index, column=4, value=start_altitude) #mark thermal beginning in session sheet in XLS
                        ws.cell(row=i, column=5, value=end_altitude) #mark thermal end in session sheet in XLS
                        session_records[session_number].thermal_count += 1
                        session_records[session_number].total_thermal_altitude_gain += altitude_gain
                        session_records[session_number].total_thermal_duration += duration
                        print("Session", session_number, "Thermal", thermal_index, "- Start alt", round(start_altitude,1), "\tEnd alt", round(end_altitude,1)) #, "start index", last_trough_bottom_index, "end index", i)
                    i=j # move starting point for next thermal peak to the start of the last thermal
                    break
                        
                elif(vario_records[j].trough_bottom == True):
                    last_event_was_a_trough = True
                    last_trough_bottom_timestamp = vario_records[j].timestamp
                    last_trough_bottom_altitude = vario_records[j].altitude_smoothed
                    last_trough_bottom_index = j

            thermal_index += 1
        i -=1
    total_thermal_duration_hms = str(datetime.timedelta(seconds=int(session_records[session_number].total_thermal_duration)))
    session_records[session_number].total_thermal_duration_hms = total_thermal_duration_hms

    if(session_records[session_number].launch_count != 0):
        session_records[session_number].thermal_launch_ratio = session_records[session_number].thermal_count / session_records[session_number].launch_count

def add_line_chart_to_ws(ws):
    from openpyxl.chart import LineChart, Reference

    chart1 = LineChart()
    chart1.title = "Session Profile"
    chart1.x_axis.title = "time (s)"
    chart1.y_axis.title = "altitude (ft)"
   # chart1.x_axis = NumericAxis(crossAx=100)
    chart1.x_axis.number_format = '#.0'
    chart1.height = 10 
    chart1.width = 30 

    # Y Axis values
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)

    chart1.series[1].graphicalProperties.line.solidFill = "00FF00"  # Green dot
    chart1.series[1].graphicalProperties.line.width = 50000  # dot thickness
    chart1.series[2].graphicalProperties.line.solidFill = "FF0000"  # Red dot
    chart1.series[2].graphicalProperties.line.width = 50000  # dot thickness

    timestamps = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart1.set_categories(timestamps)

    chart1.x_axis.title = "time (s)"
    ws.add_chart(chart1, "B3")
    

def add_thermal_bar_chart_to_ws(ws):
    from openpyxl.chart.axis import NumericAxis
    from openpyxl.chart import BarChart, Reference

    chart1 = BarChart()
    chart1.title = "Thermal Altitudes"

    # Y Axis value
    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)

    timestamps = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart1.set_categories(timestamps)    
    chart1.x_axis.title = "Thermal Number"
    chart1.y_axis.title = "alt (ft)"
    ws.add_chart(chart1, "H2")

def add_session_bar_chart_to_ws(ws):
    from openpyxl.chart.axis import NumericAxis
    from openpyxl.chart import BarChart, Reference

    chart1 = BarChart()
    chart1.title = "Launches & Thermals by Session"

    # Y Axis value
    data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)

    timestamps = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart1.set_categories(timestamps)    
    chart1.x_axis.title = "Session Number"
    chart1.y_axis.title = "Count"
    ws.add_chart(chart1, "I1")

def process_payload_packet(f, wb, thermal_records, session_records, quitProgram):
    vario_records = []
    # this routine is entered after a block of header packets has been processed
    
    # create a static variable to track session number
    if not hasattr(process_payload_packet, "session_number"):
        process_payload_packet.session_number = -1 # will start at 0 once incremented in while loop below 
        
    # Move back 4 bytes and re-read the last timestamp in case it is part of a vario record (0x40 code)
    f.seek(-4, 1)  
    timestamp_offset_hundreths_bytes = f.read(4)
    timestamp_offset_hundreths = struct.unpack('<I', timestamp_offset_hundreths_bytes)[0]
    next_byte = f.read(1)

    # if this is a 0x7C packet, get UTC time from it in the middle of the packet
    if next_byte[0] == 0x7C:
        skipped = f.read(7) # skip 7 bytes to timestamp in middle of 0x7E packet
        session_start_timestamp_unix_bytes = f.read(4)
        session_start_timestamp_unix = struct.unpack('<I', session_start_timestamp_unix_bytes)[0]
        f.read(4) #set f to start of next packet

    rowIndex =  1 #set up row index for excel sheet
    varioPacketCount = 0 # counter to track vario packets in this session
    
    while True:
        # Read next 4 bytes as time_stamp
        time_stamp_hundreths_bytes = f.read(4)

        # If EOF reached, run analytics for final session, update Session and Thermals worksheets, save XLS and quit
        if len(time_stamp_hundreths_bytes) < 4: 
            smooth_altitude_readings(vario_records)
            identify_thermal_peaks(vario_records, session_records, process_payload_packet.session_number)
            identify_launch_peaks(vario_records, session_records, process_payload_packet.session_number)
            identify_trough_bottoms(vario_records, process_payload_packet.session_number)
            identify_caught_thermals(vario_records, session_records, process_payload_packet.session_number, thermal_records, wb)
            add_line_chart_to_ws(ws)
            
            # Add rows to the Thermals ws
            ws = wb["Thermals"]
            for i in range(0, len(thermal_records)):
                rowIndex = i + 2
                ws.cell(row=rowIndex, column=1, value=i)
                ws.cell(row=rowIndex, column=2, value=thermal_records[i].session_number)
                ws.cell(row=rowIndex, column=3, value=thermal_records[i].start_altitude)
                ws.cell(row=rowIndex, column=4, value=thermal_records[i].end_altitude)
                ws.cell(row=rowIndex, column=5, value=thermal_records[i].duration)
                ws.cell(row=rowIndex, column=6, value=thermal_records[i].altitude_gain)
                ws.cell(row=rowIndex, column=7, value=thermal_records[i].avg_climb_rate)
            add_thermal_bar_chart_to_ws(ws)

            # Add rows to the Sessions ws
            ws = wb["Sessions"]
            for i in range(0, len(session_records)):
                rowIndex = i + 2
                ws.cell(row=rowIndex, column=1, value=i)
                ws.cell(row=rowIndex, column=2, value=session_records[i].start_time)
                ws.cell(row=rowIndex, column=3, value=session_records[i].session_duration_hms)
                ws["C{}".format(rowIndex)].alignment = openpyxl.styles.Alignment(horizontal='right')
                ws.cell(row=rowIndex, column=4, value=session_records[i].launch_count)
                ws.cell(row=rowIndex, column=5, value=session_records[i].thermal_count)
                ws.cell(row=rowIndex, column=6, value=session_records[i].total_thermal_altitude_gain)
                ws.cell(row=rowIndex, column=7, value=session_records[i].total_thermal_duration_hms)
                ws["G{}".format(rowIndex)].alignment = openpyxl.styles.Alignment(horizontal='right')
                ws.cell(row=rowIndex, column=8, value=round(session_records[i].thermal_launch_ratio, 1))
                # ws["H{}".format(rowIndex)].number_format =  openpyxl.styles.numbers.FORMAT_PERCENTAGE

                add_session_bar_chart_to_ws(ws)
            print("Processed", process_payload_packet.session_number + 1, "sessions,", len(thermal_records), "thermals caught.")
             
            # Update Daily Summary file
            summarize_daily_sessions(session_records)
            
            # print summary and quit
            latest_date = session_records[len(session_records)-1].start_date
            latest_date_string = latest_date.strftime("%Y_%m_%d")
            session_detail_path = os.path.splitext(tlm_file_path)[0] + "_sessions_" + latest_date_string + '.xlsx'
            wb.save(session_detail_path)
            print('Session details saved to:', session_detail_path)
            return(True) # quit program
            
        # otherwise we're not at EOF ... get the ts
        time_stamp_hundreths = struct.unpack('<I', time_stamp_hundreths_bytes)[0]
        
        # Check if we've reached another header block -- if so, we've reached the end of this session
        # Process this session
        if (time_stamp_hundreths == 0xFFFFFFFF and len(vario_records) > 0): #process  data if we just processed a vario block
            smooth_altitude_readings(vario_records)
            identify_thermal_peaks(vario_records, session_records, process_payload_packet.session_number)
            identify_launch_peaks(vario_records, session_records, process_payload_packet.session_number)
            identify_trough_bottoms(vario_records, process_payload_packet.session_number)
            identify_caught_thermals(vario_records, session_records, process_payload_packet.session_number, thermal_records, wb)
            add_line_chart_to_ws(ws)

        if (time_stamp_hundreths == 0xFFFFFFFF): # move f to next packet
            skipped = f.read(32)

            if len(skipped) < 32:
                print("Not enough bytes to skip. End of file.")
                return(False) # do not quit program
            return(False)
    
        # we've confirmed this packet is NOT a header packet or the EOF
        # read next byte....
        next_byte = f.read(1)
        if len(next_byte) < 1:
            print("End of file reached in process_payload_packet.")
            return(False) # do not quit program
        
        # ...and check if it's a vario packet (0x40)
        if next_byte[0] == 0x40:
            if(varioPacketCount % 10 == 0): # only record every 10th vario packet resulting in 1s samples
                # Check if this is the first vario packet we're writing
                if(rowIndex == 1):
                    process_payload_packet.session_number += 1
                    session_start_local_datetime = datetime.datetime.fromtimestamp(session_start_timestamp_unix)
                    session_start_local_date = session_start_local_datetime.date()
                    session_records.append(SessionRecord(session_start_local_datetime, session_start_local_date))

                    # Create new sheet
                    ws = wb.create_sheet(title=f"Ssn{process_payload_packet.session_number}")
                    # Set headers
                    headers = ["TS (s)", "Alt (ft)", "Clmb Rt (ft/min)", "Therm Start", "Therm End"]
                    for col, header in enumerate(headers, start=1):
                        ws.cell(row=1, column=col, value=header)

                    columns = ['A','B','C','D','E']
                    for column_letter in columns:
                        ws.column_dimensions[column_letter].width = 10
                        ws["{}1".format(column_letter)].font = openpyxl.styles.Font(bold=True)
                        ws["{}1".format(column_letter)].alignment = openpyxl.styles.Alignment(horizontal = 'center', wrap_text = True)
                    rowIndex += 1 # move to second row for data

                # Read rest of vario packet and convert values
                sID_bytes = f.read(1)
                altitude_bytes = f.read(2)
                delta_250ms_bytes = f.read(2) # we read this vario value, but don't do anything with it
                delta_500ms_bytes = f.read(2) # we read this vario value, but don't do anything with it
                delta_1000ms_bytes = f.read(2)
                delta_1500ms_bytes = f.read(2) # we read this vario value, but don't do anything with it
                delta_2000ms_bytes = f.read(2) # we read this vario value, but don't do anything with it
                delta_3000ms_bytes = f.read(2) # we read this vario value, but don't do anything with it
                altitude = struct.unpack('>H', altitude_bytes)[0]/10.0 * 3.28084  # Convert to feet
                delta_1000ms = struct.unpack('>h', delta_1000ms_bytes)[0]/10.0 * 3.28084 # Convert to feet
                delta_timestamp_hundreths = time_stamp_hundreths - timestamp_offset_hundreths
                delta_timestamp = delta_timestamp_hundreths / 100.0  # Convert to seconds
                
                if(altitude < 1000): # write to XLS only for valid data (altitiude under 1000 ft)
                    ws.cell(row=rowIndex, column=1, value=round(delta_timestamp, 1))
                    ws.cell(row=rowIndex, column=2, value=round(altitude, 1))
                    ws.cell(row=rowIndex, column=3, value=round(delta_1000ms, 1))
                    vario_records.append(VarioRecord(delta_timestamp, altitude, delta_1000ms))    
                    varioPacketCount += 1
                    rowIndex += 1
           
            else: # skip this vario packet -- only record every 10th packet
                varioPacketCount += 1
                f.read(15) # move f to beginning of next packet
            
        #not a vario packet, f to beginning of next packet
        else:
            skipped = f.read(15)
            if len(skipped) < 15:
                return(False) # do not quit program


'''def read_config():
    # Create a ConfigParser object
    config = configparser.ConfigParser()

    # Read the configuration file
    try:
        f = open('osprey_config.ini', 'r')
        
    except Exception as e:
        units = input("Select the units of measure.  Enter 'm' for metric, or 'i' for imperial: ")
        f= open('osprey_config.ini', 'r')
        if units == 'i' or units == 'I':
            config_string = "units = imperial"
        elif units == 'm' or units == 'M':
            config_string = "units = metric"
        else:
            print("Invalid units entered, using metric.")
        f.write("[General]\n")
        f.write(config_string)

    units = config.get('General', 'units')
        config_values = {
        'units': units
    }
    
    # Return a dictionary with the retrieved values
    return config_values '''


# if __name__ == "__main__":
    # Call the function to read the configuration file
    # config_data = read_config()

    # Print the retrieved values
    # print("Debug Mode:", config_data['debug_mode'])
    # print("Log Level:", config_data['log_level'])
    # print("Database Name:", config_data['db_name'])
    # print("Database Host:", config_data['db_host'])
    # print("Database Port:", config_data['db_port'])

def main():
    thermal_records = []
    session_records = []
    quitProgram = False

    source_dir = os.path.abspath(__file__)

    # Ask user for input file
    # filename = input("Enter the TLM file path: ")
    global tlm_file_path 
    tlm_file_path = "/Users/tayloranderson/dev/spektrum/Avanti.TLM"

    # tlm_file_path = filedialog.askopenfilename(
    #     initialdir=source_dir,
    #     title="Select Spektrum TLM file",
    #     filetypes=[("Spektrum log data", "*.tlm"), ("All files", "*.*")]
    # )
  
    if not os.path.exists(tlm_file_path):
        print(f"Error: File '{tlm_file_path}' not found.")
        return(False)
    
    try:
        f = open(tlm_file_path, 'rb')
    
    except Exception as e:
        print(f"Error reading file: {e}")
        return(False)

    # Set up Sessions worksheet now so we can populate it later
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws.title = "Sessions"
    headers = ['Session #', 'Session Start', 'Session Duration', 'Launches', 'Thermals', 'Total Thm Alt Gain (ft)', 'Total Thm Duration', 'Thermal/Lauch Ratio']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    columns = ['A','B','C','D','E','F','G', 'H']
    for column_letter in columns:
        ws["{}1".format(column_letter)].font = openpyxl.styles.Font(bold=True)
        ws["{}1".format(column_letter)].alignment = openpyxl.styles.Alignment(horizontal = 'center', wrap_text = True)
    #    horizontal='center', 
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

     # Also set up Thermals worksheet
    ws = wb.create_sheet(title=f"Thermals")
    headers = ['Thermal #', 'Session #', 'Start Alt (ft)', 
                'End Alt (ft)', 'Therm Dur (s)', 'Alt Gain (ft)', 'Avg Clb Rt (ft/s)']
    for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

    columns = ['A','B','C','D','E','F','G']
    for column_letter in columns:
        ws.column_dimensions[column_letter].width = 10
        ws["{}1".format(column_letter)].font = openpyxl.styles.Font(bold=True)
        ws["{}1".format(column_letter)].alignment = openpyxl.styles.Alignment(horizontal = 'center', wrap_text = True)

    while (quitProgram == False):
        process_header_packets(f)
        quitProgram = process_payload_packet(f, wb, thermal_records, session_records, quitProgram)

if __name__ == "__main__":
    main()
